# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import sys
reload(sys)
sys.setdefaultencoding('utf8')
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from scipy.misc import imread
# 中文情况下，只有通过分词才能产生词云
import jieba
# import jieba.analyse as analyse
import os

from openpyxl import load_workbook

class FilmcloudPipeline(object):
    def process_item(self, item, spider):

        print item

        # 分词
        cut_text = " ".join(jieba.cut(item['summary']))
        # cut_text = " ".join(analyse.extract_tags(item['summary'], topK=20, withWeight=False))

        cut_text += " " + item["name"] + " " + str(item["score"]) + " "+ item["director"]
        font = r'SourceHanSansCN-ExtraLight.otf'
        wordcloud = WordCloud(background_color='black', scale=1.5, font_path=font, width=1000, height=1000,
                              max_words=200).generate(cut_text)

        # plt.imshow(wordcloud)
        # plt.axis('off')
        # plt.show()
        # 保存图片
        if not os.path.exists(os.curdir + "/cloud_image/"):
            os.mkdir(os.curdir + "/cloud_image/")
        path = os.curdir+"/cloud_image/"
        image_cloud_fileName = u'%s_cloud.jpg'%(item["name"])
        image_cloud_fileName = path + image_cloud_fileName
        print image_cloud_fileName
        print wordcloud.to_file(image_cloud_fileName)

        # wb = load_workbook("file.xlsx")
        # sheet = wb.index(0)
        #
        # # 通过get_sheet()获取的sheet有write()方法
        # # sheet_name = data.sheet_names()
        # # table = data.get_sheet(0)
        #
        # nrows = sheet.max_row
        # ncols = sheet.max_column
        #
        # s
        #
        # img_path = os.curdir + "/moive_image/"
        #
        # for i in range(ncols):
        #     if i == 0:
        #         table.write(nrows+1, i, item["name"])
        #     elif i == 1:
        #         table.insert_image(nrows+1, i, img_path + item["picture"])
        #     elif i == 2:
        #         table.write(nrows+1, i, item["director"])
        #     elif i == 3:
        #         table.write(nrows+1, i, item["summary"])
        #     else:
        #         table.insert_image(nrows + 1, i, image_cloud_fileName)
        #
        # table.save()

        return item
